from pathlib import Path
import os
import sys

import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text
from sqlalchemy.engine import URL


def get_database_engine():
    load_dotenv()

    required_variables = [
        "MYSQL_HOST",
        "MYSQL_PORT",
        "MYSQL_USER",
        "MYSQL_PASSWORD",
        "MYSQL_DATABASE",
    ]

    missing_variables = [
        variable for variable in required_variables
        if not os.getenv(variable)
    ]

    if missing_variables:
        raise ValueError(
            f"Missing environment variables: {', '.join(missing_variables)}"
        )

    database_url = URL.create(
        drivername="mysql+pymysql",
        username=os.getenv("MYSQL_USER"),
        password=os.getenv("MYSQL_PASSWORD"),
        host=os.getenv("MYSQL_HOST"),
        port=int(os.getenv("MYSQL_PORT", "3306")),
        database=os.getenv("MYSQL_DATABASE"),
    )

    return create_engine(database_url, pool_pre_ping=True)


def load_overall_kpis(engine) -> pd.DataFrame:
    query = text(
        """
        SELECT
            SUM(units_produced) AS total_units_produced,
            ROUND(
                SUM(good_units) /
                NULLIF(SUM(units_produced), 0) * 100,
                2
            ) AS yield_rate_pct,
            ROUND(
                SUM(defective_units) /
                NULLIF(SUM(units_produced), 0) * 100,
                2
            ) AS defect_rate_pct,
            ROUND(
                SUM(units_produced) /
                NULLIF(SUM(planned_capacity), 0) * 100,
                2
            ) AS capacity_utilization_pct,
            SUM(abnormal_flag) AS abnormal_batches
        FROM production_records;
        """
    )

    return pd.read_sql(query, engine)


def load_daily_kpis(engine) -> pd.DataFrame:
    query = text(
        """
        SELECT
            production_date,
            SUM(units_produced) AS units_produced,
            ROUND(
                SUM(defective_units) /
                NULLIF(SUM(units_produced), 0) * 100,
                2
            ) AS defect_rate_pct,
            ROUND(
                SUM(good_units) /
                NULLIF(SUM(units_produced), 0) * 100,
                2
            ) AS yield_rate_pct,
            ROUND(
                SUM(units_produced) /
                NULLIF(SUM(planned_capacity), 0) * 100,
                2
            ) AS capacity_utilization_pct,
            SUM(abnormal_flag) AS abnormal_batches
        FROM production_records
        GROUP BY production_date
        ORDER BY production_date;
        """
    )

    return pd.read_sql(query, engine)


def load_machine_kpis(engine) -> pd.DataFrame:
    query = text(
        """
        SELECT
            machine_id,
            SUM(units_produced) AS units_produced,
            ROUND(
                SUM(defective_units) /
                NULLIF(SUM(units_produced), 0) * 100,
                2
            ) AS defect_rate_pct,
            ROUND(
                AVG(downtime_minutes),
                2
            ) AS avg_downtime_minutes,
            ROUND(
                SUM(units_produced) /
                NULLIF(SUM(planned_capacity), 0) * 100,
                2
            ) AS capacity_utilization_pct,
            SUM(abnormal_flag) AS abnormal_batches
        FROM production_records
        GROUP BY machine_id
        ORDER BY machine_id;
        """
    )

    return pd.read_sql(query, engine)


def load_shift_kpis(engine) -> pd.DataFrame:
    query = text(
        """
        SELECT
            shift_name,
            SUM(units_produced) AS units_produced,
            ROUND(
                SUM(good_units) /
                NULLIF(SUM(units_produced), 0) * 100,
                2
            ) AS yield_rate_pct,
            ROUND(
                SUM(defective_units) /
                NULLIF(SUM(units_produced), 0) * 100,
                2
            ) AS defect_rate_pct,
            ROUND(
                AVG(downtime_minutes),
                2
            ) AS avg_downtime_minutes
        FROM production_records
        GROUP BY shift_name
        ORDER BY shift_name;
        """
    )

    return pd.read_sql(query, engine)


def write_dataframe(worksheet, dataframe, start_row=1, start_column=1):
    for column_offset, column_name in enumerate(
        dataframe.columns,
        start=start_column,
    ):
        worksheet.cell(
            row=start_row,
            column=column_offset,
            value=column_name,
        )

    for row_offset, row_values in enumerate(
        dataframe.itertuples(index=False),
        start=start_row + 1,
    ):
        for column_offset, value in enumerate(
            row_values,
            start=start_column,
        ):
            worksheet.cell(
                row=row_offset,
                column=column_offset,
                value=value,
            )


def style_data_sheet(worksheet):
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = 0

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(
            max_length + 2,
            24,
        )


def add_kpi_card(
    worksheet,
    label_cell,
    value_cell,
    label,
    value,
    number_format,
):
    worksheet[label_cell] = label
    worksheet[value_cell] = value

    worksheet[label_cell].font = Font(
        bold=True,
        size=11,
    )
    worksheet[value_cell].font = Font(
        bold=True,
        size=18,
    )

    worksheet[label_cell].alignment = Alignment(
        horizontal="center",
    )
    worksheet[value_cell].alignment = Alignment(
        horizontal="center",
    )

    worksheet[value_cell].number_format = number_format


def create_dashboard(
    overall_kpis,
    daily_kpis,
    machine_kpis,
    shift_kpis,
    output_path,
):
    workbook = Workbook()

    dashboard = workbook.active
    dashboard.title = "Dashboard"

    daily_sheet = workbook.create_sheet("Daily KPI Data")
    machine_sheet = workbook.create_sheet("Machine KPI Data")
    shift_sheet = workbook.create_sheet("Shift KPI Data")

    write_dataframe(daily_sheet, daily_kpis)
    write_dataframe(machine_sheet, machine_kpis)
    write_dataframe(shift_sheet, shift_kpis)

    style_data_sheet(daily_sheet)
    style_data_sheet(machine_sheet)
    style_data_sheet(shift_sheet)

    dashboard.merge_cells("A1:L2")
    dashboard["A1"] = "Manufacturing KPI Dashboard"
    dashboard["A1"].font = Font(
        bold=True,
        size=22,
    )
    dashboard["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    overall = overall_kpis.iloc[0]

    add_kpi_card(
        dashboard,
        "A4",
        "A5",
        "Total Units Produced",
        int(overall["total_units_produced"]),
        "#,##0",
    )

    add_kpi_card(
        dashboard,
        "C4",
        "C5",
        "Yield Rate",
        float(overall["yield_rate_pct"]) / 100,
        "0.00%",
    )

    add_kpi_card(
        dashboard,
        "E4",
        "E5",
        "Defect Rate",
        float(overall["defect_rate_pct"]) / 100,
        "0.00%",
    )

    add_kpi_card(
        dashboard,
        "G4",
        "G5",
        "Capacity Utilization",
        float(overall["capacity_utilization_pct"]) / 100,
        "0.00%",
    )

    add_kpi_card(
        dashboard,
        "I4",
        "I5",
        "Abnormal Batches",
        int(overall["abnormal_batches"]),
        "#,##0",
    )

    daily_rows = len(daily_kpis) + 1
    machine_rows = len(machine_kpis) + 1
    shift_rows = len(shift_kpis) + 1

    production_chart = LineChart()
    production_chart.title = "Daily Production Output"
    production_chart.y_axis.title = "Units Produced"
    production_chart.x_axis.title = "Production Date"
    production_chart.height = 8
    production_chart.width = 15

    production_data = Reference(
        daily_sheet,
        min_col=2,
        min_row=1,
        max_row=daily_rows,
    )

    production_dates = Reference(
        daily_sheet,
        min_col=1,
        min_row=2,
        max_row=daily_rows,
    )

    production_chart.add_data(
        production_data,
        titles_from_data=True,
    )
    production_chart.set_categories(production_dates)

    dashboard.add_chart(production_chart, "A8")

    defect_chart = LineChart()
    defect_chart.title = "Daily Defect Rate"
    defect_chart.y_axis.title = "Defect Rate (%)"
    defect_chart.x_axis.title = "Production Date"
    defect_chart.height = 8
    defect_chart.width = 15

    defect_data = Reference(
        daily_sheet,
        min_col=3,
        min_row=1,
        max_row=daily_rows,
    )

    defect_chart.add_data(
        defect_data,
        titles_from_data=True,
    )
    defect_chart.set_categories(production_dates)

    dashboard.add_chart(defect_chart, "H8")

    machine_defect_chart = BarChart()
    machine_defect_chart.title = "Defect Rate by Machine"
    machine_defect_chart.y_axis.title = "Defect Rate (%)"
    machine_defect_chart.x_axis.title = "Machine"
    machine_defect_chart.height = 8
    machine_defect_chart.width = 15

    machine_defect_data = Reference(
        machine_sheet,
        min_col=3,
        min_row=1,
        max_row=machine_rows,
    )

    machine_names = Reference(
        machine_sheet,
        min_col=1,
        min_row=2,
        max_row=machine_rows,
    )

    machine_defect_chart.add_data(
        machine_defect_data,
        titles_from_data=True,
    )
    machine_defect_chart.set_categories(machine_names)

    dashboard.add_chart(machine_defect_chart, "A24")

    downtime_chart = BarChart()
    downtime_chart.title = "Average Downtime by Machine"
    downtime_chart.y_axis.title = "Minutes"
    downtime_chart.x_axis.title = "Machine"
    downtime_chart.height = 8
    downtime_chart.width = 15

    downtime_data = Reference(
        machine_sheet,
        min_col=4,
        min_row=1,
        max_row=machine_rows,
    )

    downtime_chart.add_data(
        downtime_data,
        titles_from_data=True,
    )
    downtime_chart.set_categories(machine_names)

    dashboard.add_chart(downtime_chart, "H24")

    shift_chart = BarChart()
    shift_chart.title = "Production Output by Shift"
    shift_chart.y_axis.title = "Units Produced"
    shift_chart.x_axis.title = "Shift"
    shift_chart.height = 8
    shift_chart.width = 15

    shift_data = Reference(
        shift_sheet,
        min_col=2,
        min_row=1,
        max_row=shift_rows,
    )

    shift_names = Reference(
        shift_sheet,
        min_col=1,
        min_row=2,
        max_row=shift_rows,
    )

    shift_chart.add_data(
        shift_data,
        titles_from_data=True,
    )
    shift_chart.set_categories(shift_names)

    dashboard.add_chart(shift_chart, "A40")

    for column in range(1, 13):
        dashboard.column_dimensions[
            get_column_letter(column)
        ].width = 14

    dashboard.sheet_view.showGridLines = False
    dashboard.freeze_panes = "A3"

    workbook.save(output_path)


def main():
    project_root = Path(__file__).resolve().parents[1]
    reports_directory = project_root / "reports"
    reports_directory.mkdir(parents=True, exist_ok=True)

    output_path = (
        reports_directory /
        "manufacturing_dashboard.xlsx"
    )

    try:
        engine = get_database_engine()

        overall_kpis = load_overall_kpis(engine)
        daily_kpis = load_daily_kpis(engine)
        machine_kpis = load_machine_kpis(engine)
        shift_kpis = load_shift_kpis(engine)

        create_dashboard(
            overall_kpis,
            daily_kpis,
            machine_kpis,
            shift_kpis,
            output_path,
        )

        print("Manufacturing dashboard created successfully.")
        print(f"Saved dashboard to: {output_path}")

    except Exception as error:
        print(
            f"Dashboard creation failed: {error}",
            file=sys.stderr,
        )
        sys.exit(1)


if __name__ == "__main__":
    main()